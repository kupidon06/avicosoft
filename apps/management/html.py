import os
import json
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook

from django.conf import settings
from django.contrib import messages
from django.utils import timezone
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.db.models import Q,Sum
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from .models import Building, Breed, Batch, Feed, Treatment, DailyLog, Feeding, EggCollection, TreatmentHistory, Provision, ExpenseCategory, Expense,StockLoss
from apps.ventes.models import BatchAllocation,OrderItem
from .forms import BuildingForm, BreedForm, BatchForm, FeedForm, TreatmentForm, DailyLogForm, FeedingForm, EggCollectionForm, TreatmentHistoryForm, ProvisionForm, ExpenseCategoryForm, ExpenseForm,StockLossForm
from apps.users.models import Fournisseur
from apps.profile.models import CompanyProfile
from apps.users.decorators import role_required

login_path = '/accounts/login'

# Vues pour les Bâtiments
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee']) 
def building_list(request):
    query = request.GET.get('q')
    if query:
        buildings = Building.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        buildings = Building.objects.all().order_by('id')

    paginator = Paginator(buildings, 10)  # Show 10 buildings per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = BuildingForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('building_list')
    else:
        form = BuildingForm()

    context = {
        'buildings': page_obj,
        'form': form,
    }

    return render(request, 'apps/buildings/list.html', context)

@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee']) 
def building_list_print(request):
    query = request.GET.get('q')
    if query:
        buildings = Building.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        buildings = Building.objects.all().order_by('id')

    company = CompanyProfile.objects.all().first()
    context = {
        'buildings': buildings,
        'company':company
    }

    return render(request, 'apps/buildings/list_print.html', context)


@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee']) 
def building_list_export(request):
    query = request.GET.get('q')
    if query:
        buildings = Building.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        buildings = Building.objects.all().order_by('id')

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Buildings"

    # Ajouter les en-têtes
    ws.append(["ID", "Name", "Capacity", "Details", "Photo URL"])

    # Ajouter les données des bâtiments
    for building in buildings:
        ws.append([building.id, building.name, building.capacity, building.details, (settings.SITE_URL + building.photo.url) if building.photo else ""])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=buildings.xlsx'

    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def update_building(request, id):
    building = get_object_or_404(Building, id=id)

    if request.method == "POST":
        building.name = request.POST.get('name', building.name)
        building.details = request.POST.get('details', building.details)
        building.capacity = request.POST.get('capacity', building.capacity)

        if 'image' in request.FILES:
            if building.photo:
                old_image_path = os.path.join(settings.MEDIA_ROOT, str(building.photo))
                if os.path.isfile(old_image_path):
                    try:
                        os.remove(old_image_path)
                    except Exception as e:
                        print(f"Error removing old image: {e}")
                else:
                    print(f"Old image not found: {old_image_path}")

            building.photo = request.FILES['image']

        try:
            building.save()
            print("Building saved successfully")
        except Exception as e:
            print(f"Error saving building: {e}")

        return redirect('building_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def delete_building(request, id):
    building = get_object_or_404(Building, id=id)
    if building.photo:
        old_image_path = os.path.join(settings.MEDIA_ROOT, str(building.photo))
        if os.path.isfile(old_image_path):
            os.remove(old_image_path)
    building.delete()
    return redirect('building_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def bulk_delete(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            building_ids = data.get('building_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not building_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucun bâtiment sélectionné.'}, status=400)

            elif action == 'delete':
                # Supprimer les bâtiments et leurs images
                buildings = Building.objects.filter(id__in=building_ids)
                for building in buildings:
                    if building.photo:
                        old_image_path = os.path.join(settings.MEDIA_ROOT, str(building.photo))
                        if os.path.isfile(old_image_path):
                            os.remove(old_image_path)
                buildings_deleted, _ = buildings.delete()

                return JsonResponse({'status': 'success', 'message': f'{buildings_deleted} bâtiment(s) supprimé(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Vues pour les Races
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee']) 
def breed_list(request):
    query = request.GET.get('q')
    if query:
        breeds = Breed.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        breeds = Breed.objects.all()

    paginator = Paginator(breeds, 10)  # Show 10 breeds per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = BreedForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('breed_list')
    else:
        form = BreedForm()

    context = {
        'breeds': page_obj,
        'form': form,
    }

    return render(request, 'apps/breeds/list.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee']) 
def breed_list_print(request):
    query = request.GET.get('q')
    if query:
        breeds = Breed.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        breeds = Breed.objects.all().order_by('id')
    company = CompanyProfile.objects.all().first()
    context = {
        'breeds': breeds,
        'company':company
    }

    return render(request, 'apps/breeds/list_print.html', context)


@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def breed_list_export(request):
    query = request.GET.get('q')
    if query:
        breeds = Breed.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        breeds = Breed.objects.all().order_by('id')

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Breeds"

    # Ajouter les en-têtes
    ws.append(["ID", "Name", "Details", "Photo URL"])

    # Ajouter les données des races
    for breed in breeds:
        ws.append([breed.id, breed.name, breed.details, (settings.SITE_URL + breed.photo.url) if breed.photo else ""])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=breeds.xlsx'

    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def update_breed(request, id):
    breed = get_object_or_404(Breed, id=id)

    if request.method == "POST":
        breed.name = request.POST.get('name', breed.name)
        breed.details = request.POST.get('details', breed.details)

        if 'image' in request.FILES:
            if breed.photo:
                old_image_path = os.path.join(settings.MEDIA_ROOT, str(breed.photo))
                if os.path.isfile(old_image_path):
                    try:
                        os.remove(old_image_path)
                    except Exception as e:
                        print(f"Error removing old image: {e}")
                else:
                    print(f"Old image not found: {old_image_path}")

            breed.photo = request.FILES['image']

        try:
            breed.save()
            print("Breed saved successfully")
        except Exception as e:
            print(f"Error saving breed: {e}")

        return redirect('breed_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def delete_breed(request, id):
    breed = get_object_or_404(Breed, id=id)
    if breed.photo:
        old_image_path = os.path.join(settings.MEDIA_ROOT, str(breed.photo))
        if os.path.isfile(old_image_path):
            os.remove(old_image_path)
    breed.delete()
    return redirect('breed_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def bulk_delete_breed(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            breed_ids = data.get('breed_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not breed_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucune race sélectionnée.'}, status=400)

            elif action == 'delete':
                # Supprimer les races et leurs images
                breeds = Breed.objects.filter(id__in=breed_ids)
                for breed in breeds:
                    if breed.photo:
                        old_image_path = os.path.join(settings.MEDIA_ROOT, str(breed.photo))
                        if os.path.isfile(old_image_path):
                            os.remove(old_image_path)
                breeds_deleted, _ = breeds.delete()

                return JsonResponse({'status': 'success', 'message': f'{breeds_deleted} race(s) supprimé(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Vues pour les Batches
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee']) 
def batch_list(request):
    query = request.GET.get('q')
    if query:
        batches = Batch.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        batches = Batch.objects.all().order_by('id')

    for batch in batches:
        if batch.arrival_date:
            batch.arrival_date = batch.arrival_date.strftime('%Y-%m-%d')
    paginator = Paginator(batches, 10)  # Show 10 batches per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = BatchForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('batch_list')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = BatchForm()
    breeds = Breed.objects.all()
    buildings = Building.objects.all()
    STATUS_CHOICES = [
        ('active', 'Active'),
        ('inactive', 'Inactive'),
        ('quarantine', 'Quarantine'),
    ]

    context = {
        'batches': page_obj,
        'form': form,
        'breeds': breeds,
        'buildings': buildings,
        'STATUS_CHOICES': STATUS_CHOICES,
    }

    return render(request, 'apps/batches/list.html', context)

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def update_batch(request, id):
    batch = get_object_or_404(Batch, id=id)

    if request.method == "POST":
        batch.name = request.POST.get('name', batch.name)
        batch.breed = get_object_or_404(Breed, id=request.POST.get('breed'))
        batch.building = get_object_or_404(Building, id=request.POST.get('building'))
        batch.arrival_date = request.POST.get('arrival_date', batch.arrival_date)
        batch.arrival_age = request.POST.get('arrival_age', batch.arrival_age)
        batch.arrival_quantity = request.POST.get('arrival_quantity', batch.arrival_quantity)
        batch.status = request.POST.get('status', batch.status)
        batch.details = request.POST.get('details', batch.details)

        batch.save()
        return redirect('batch_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def delete_batch(request, id):
    batch = get_object_or_404(Batch, id=id)
    batch.delete()
    return redirect('batch_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def bulk_delete_batch(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            batch_ids = data.get('batch_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not batch_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucun batch sélectionné.'}, status=400)

            elif action == 'delete':
                # Supprimer les batches
                batches = Batch.objects.filter(id__in=batch_ids)
                batches_deleted, _ = batches.delete()

                return JsonResponse({'status': 'success', 'message': f'{batches_deleted} batch(s) supprimé(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Vues pour les Aliments
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def feed_list(request):
    query = request.GET.get('q')
    if query:
        feeds = Feed.objects.filter(Q(name__icontains=query)).distinct()
    else:
        feeds = Feed.objects.all()

    paginator = Paginator(feeds, 10)  # Show 10 feeds per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = FeedForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('feed_list')
    else:
        form = FeedForm()

    context = {
        'feeds': page_obj,
        'form': form,
    }

    return render(request, 'apps/feeds/list.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def feed_list_print(request):
    query = request.GET.get('q')
    if query:
        feeds = Feed.objects.filter(Q(name__icontains=query)).distinct()
    else:
        feeds = Feed.objects.all().order_by('id')
    company = CompanyProfile.objects.all().first()

    context = {
        'feeds': feeds,
        'company':company
    }

    return render(request, 'apps/feeds/list_print.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def feed_list_export(request):
    query = request.GET.get('q')
    if query:
        feeds = Feed.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        feeds = Feed.objects.all().order_by('id')

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Feeds"

    # Ajouter les en-têtes
    ws.append(["Name", "Details", "Unit Price", "Unit Measure"])

    # Ajouter les données des aliments
    for feed in feeds:
        ws.append([feed.name, feed.details, feed.unit_price, feed.unit_measure])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=feeds.xlsx'

    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def update_feed(request, id):
    feed = get_object_or_404(Feed, id=id)

    if request.method == "POST":
        feed.name = request.POST.get('name', feed.name)
        feed.details = request.POST.get('details', feed.details)
        feed.unit_price = request.POST.get('unit_price', feed.unit_price)
        feed.unit_measure = request.POST.get('unit_measure', feed.unit_measure)

        feed.save()
        return redirect('feed_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def delete_feed(request, id):
    feed = get_object_or_404(Feed, id=id)
    feed.delete()
    return redirect('feed_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def bulk_delete_feed(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            feed_ids = data.get('feed_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not feed_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucun aliment sélectionné.'}, status=400)

            elif action == 'delete':
                # Supprimer les aliments
                feeds = Feed.objects.filter(id__in=feed_ids)
                feeds_deleted, _ = feeds.delete()

                return JsonResponse({'status': 'success', 'message': f'{feeds_deleted} aliment(s) supprimé(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Vues pour les Traitements
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def treatment_list(request):
    query = request.GET.get('q')
    if query:
        treatments = Treatment.objects.filter(Q(name__icontains=query)).distinct()
    else:
        treatments = Treatment.objects.all()

    paginator = Paginator(treatments, 10)  # Show 10 treatments per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = TreatmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('treatment_list')
    else:
        form = TreatmentForm()

    context = {
        'treatments': page_obj,
        'form': form,
    }

    return render(request, 'apps/treatments/list.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def treatment_list_print(request):
    query = request.GET.get('q')
    if query:
        treatments = Treatment.objects.filter(Q(name__icontains=query)).distinct()
    else:
        treatments = Treatment.objects.all().order_by('id')
    company = CompanyProfile.objects.all().first()
    context = {
        'treatments': treatments,
        'company':company
    }

    return render(request, 'apps/treatments/list_print.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def treatment_list_export(request):
    query = request.GET.get('q')
    if query:
        treatments = Treatment.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        treatments = Treatment.objects.all().order_by('id')

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Treatments"

    # Ajouter les en-têtes
    ws.append(["Name", "Details", "Duration Days"])

    # Ajouter les données des traitements
    for treatment in treatments:
        ws.append([treatment.name, treatment.details, treatment.duration_days])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=treatments.xlsx'

    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def update_treatment(request, id):
    treatment = get_object_or_404(Treatment, id=id)

    if request.method == "POST":
        treatment.name = request.POST.get('name', treatment.name)
        treatment.details = request.POST.get('details', treatment.details)
        treatment.duration_days = request.POST.get('duration_days', treatment.duration_days)

        treatment.save()
        return redirect('treatment_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def delete_treatment(request, id):
    treatment = get_object_or_404(Treatment, id=id)
    treatment.delete()
    return redirect('treatment_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def bulk_delete_treatment(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            treatment_ids = data.get('treatment_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not treatment_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucun traitement sélectionné.'}, status=400)

            elif action == 'delete':
                # Supprimer les traitements
                treatments = Treatment.objects.filter(id__in=treatment_ids)
                treatments_deleted, _ = treatments.delete()

                return JsonResponse({'status': 'success', 'message': f'{treatments_deleted} traitement(s) supprimé(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Vues pour les Journaux Quotidiens

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def daily_log_list(request, batch_id):
    daily_logs = DailyLog.objects.filter(batch_id=batch_id).order_by('log_date', 'id')
    query = request.GET.get('q')

    if query:
        try:
            if 'to' in query:
                start_date, end_date = query.split('to')
                start_date = datetime.strptime(start_date.strip(), '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date.strip(), '%Y-%m-%d').date()
                daily_logs = daily_logs.filter(log_date__range=(start_date, end_date))
            else:
                query_date = datetime.strptime(query.strip(), '%Y-%m-%d').date()
                daily_logs = daily_logs.filter(log_date=query_date)
        except ValueError:
            daily_logs = DailyLog.objects.none()
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")

    grouped_logs = defaultdict(list)
    for daily_log in daily_logs:
        if daily_log.log_date:
            log_date_str = daily_log.log_date.strftime('%Y-%m-%d')
            daily_log.log_date = daily_log.log_date.strftime('%Y-%m-%d')
            grouped_logs[log_date_str].append(daily_log)

    paginator = Paginator(list(grouped_logs.items()), 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    batch = get_object_or_404(Batch, id=batch_id)

    if request.method == 'POST':
        form = DailyLogForm(request.POST)
        if form.is_valid():
            daily_log = form.save(commit=False)
            daily_log.batch_id = batch_id  # Associe automatiquement le batch avant validation

            # Validation manuelle
            if daily_log.log_date > timezone.now().date():
                messages.error(request, "The log date cannot be in the future.")
            elif daily_log.log_date < batch.arrival_date:
                messages.error(request, "The log date cannot be before the batch arrival date.")
            elif daily_log.deceased_quantity < 0 or daily_log.sick_quantity < 0 or daily_log.living_quantity < 0:
                messages.error(request, "Quantities cannot be negative.")
            else:
                # Vérification des quantités par rapport au batch
                total_deceased = (
                    batch.dailylog_set.exclude(id=daily_log.id)
                    .aggregate(total=Sum('deceased_quantity'))['total'] or 0
                ) + daily_log.deceased_quantity

                if total_deceased > batch.arrival_quantity:
                    messages.error(request, f"Total deceased quantity ({total_deceased}) exceeds the batch arrival quantity ({batch.arrival_quantity}).")
                else:
                    # Vérification de la cohérence des vivants
                    expected_living = batch.arrival_quantity - total_deceased
                    if daily_log.living_quantity != expected_living:
                        messages.error(request, f"Living quantity ({daily_log.living_quantity}) does not match the expected value ({expected_living}).")
                    else:
                        # Si toutes les validations sont passées, sauvegarder l'enregistrement
                        daily_log.save()
                        messages.success(request, "Daily log added successfully.")
                        return redirect('daily_log_list', batch_id=batch_id)

        else:
            messages.error(request, "There are some errors in the submitted data. Please check and try again.")

    form = DailyLogForm()
    context = {
        'grouped_logs': page_obj,
        'batch': batch,
        'form': form,
    }

    return render(request, 'apps/daily_logs/list.html', context)



@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def daily_log_list_print(request, batch_id):
    daily_logs = DailyLog.objects.filter(batch_id=batch_id).order_by('log_date', 'id')
    query = request.GET.get('q')
    if query:
        daily_logs = daily_logs.filter(Q(details__icontains=query)).distinct().order_by('id')
    else:
        daily_logs = daily_logs
    company = CompanyProfile.objects.all().first()

    context = {
        'daily_logs': daily_logs,
        'company':company
    }

    return render(request, 'apps/daily_logs/list_print.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def daily_log_list_export(request, batch_id):
    daily_logs = DailyLog.objects.filter(batch_id=batch_id).order_by('log_date', 'id')
    query = request.GET.get('q')
    if query:
        daily_logs = daily_logs.filter(Q(details__icontains=query)).distinct().order_by('id')
    else:
        daily_logs = daily_logs

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "DailyLogs"

    # Ajouter les en-têtes
    ws.append(["Date", "Batch", "Living Quantity", "Deceased Quantity", "Sick Quantity", "Details"])

    # Ajouter les données des journaux quotidiens
    for log in daily_logs:
        ws.append([log.log_date, log.batch.name, log.living_quantity, log.deceased_quantity, log.sick_quantity, log.details])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=DailyLogs.xlsx'

    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response



from django.utils.dateparse import parse_date

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def update_daily_log(request, id):
    daily_log = get_object_or_404(DailyLog, id=id)
    batch = daily_log.batch
    batch_id = batch.id

    if request.method == "POST":
        # Récupérer et valider la date
        log_date_str = request.POST.get('log_date')
        log_date = parse_date(log_date_str)  # Parse_date retourne None si le format est invalide

        if not log_date:
            messages.error(request, "Invalid log date format. Please use YYYY-MM-DD.")
            return redirect('daily_log_list', batch_id=batch_id)

        # Récupérer les quantités et les vérifier
        try:
            living_quantity = int(request.POST.get('living_quantity', 0))
            deceased_quantity = int(request.POST.get('deceased_quantity', 0))
            sick_quantity = int(request.POST.get('sick_quantity', 0))
        except ValueError:
            messages.error(request, "Quantities must be valid integers.")
            return redirect('daily_log_list', batch_id=batch_id)

        details = request.POST.get('details', '')

        daily_log.log_date = log_date
        daily_log.living_quantity = living_quantity
        daily_log.deceased_quantity = deceased_quantity
        daily_log.sick_quantity = sick_quantity
        daily_log.details = details

        # Validation manuelle des données
        if log_date > timezone.now().date():
            messages.error(request, "The log date cannot be in the future.")
        elif log_date < batch.arrival_date:
            messages.error(request, "The log date cannot be before the batch arrival date.")
        elif living_quantity < 0 or deceased_quantity < 0 or sick_quantity < 0:
            messages.error(request, "Quantities cannot be negative.")
        else:
            # Calculer les décès totaux, en excluant le log actuel
            total_deceased = (
                batch.dailylog_set.exclude(id=daily_log.id)
                .aggregate(total=Sum('deceased_quantity'))['total'] or 0
            ) + deceased_quantity

            if total_deceased > batch.arrival_quantity:
                messages.error(request, f"Total deceased quantity ({total_deceased}) exceeds the batch arrival quantity ({batch.arrival_quantity}).")
            else:
                # Vérification cohérence des vivants
                expected_living = batch.arrival_quantity - total_deceased
                if living_quantity != expected_living:
                    messages.error(request, f"Living quantity ({living_quantity}) does not match the expected value ({expected_living}).")
                else:
                    # Si tout est valide, sauvegarder et rediriger
                    daily_log.save()
                    messages.success(request, "Daily log updated successfully.")
                    return redirect('daily_log_list', batch_id=batch_id)

  
    return redirect('daily_log_list', batch_id=batch_id)




@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def delete_daily_log(request, id):
    daily_log = get_object_or_404(DailyLog, id=id)
    batch_id = daily_log.batch.id
    daily_log.delete()
    return redirect('daily_log_list', batch_id=batch_id)

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def bulk_delete_daily_log(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            dailylog_ids = data.get('dailylog_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not dailylog_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucun journal quotidien sélectionné.'}, status=400)

            elif action == 'delete':
                # Supprimer les journaux quotidiens
                dailylogs = DailyLog.objects.filter(id__in=dailylog_ids)
                dailylogs_deleted, _ = dailylogs.delete()

                return JsonResponse({'status': 'success', 'message': f'{dailylogs_deleted} journal(s) quotidien(s) supprimé(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Vues pour les Alimentations

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def feeding_list(request, batch_id):
    feedings = Feeding.objects.filter(batch_id=batch_id).order_by('feeding_date', 'id')
    query = request.GET.get('q')
    if query:
        try:
            if 'to' in query:
                start_date, end_date = query.split('to')
                start_date = datetime.strptime(start_date.strip(), '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date.strip(), '%Y-%m-%d').date()
                feedings = feedings.filter(feeding_date__range=(start_date, end_date))
            else:
                query_date = datetime.strptime(query.strip(), '%Y-%m-%d').date()
                feedings = feedings.filter(feeding_date=query_date)
        except ValueError:
            feedings = Feeding.objects.none()
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")

    grouped_feedings = defaultdict(list)
    for feeding in feedings:
        if feeding.feeding_date:
            feeding_date_str = feeding.feeding_date.strftime('%Y-%m-%d')
            feeding.feeding_date = feeding_date_str
            grouped_feedings[feeding_date_str].append(feeding)

    paginator = Paginator(list(grouped_feedings.items()), 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = FeedingForm(request.POST)
        if form.is_valid():
            feeding = form.save(commit=False)
            feeding.batch_id = batch_id

            # Vérification des stocks disponibles pour le feed utilisé
            feed = feeding.feed_type
            total_provisions = Provision.objects.filter(feed=feed).aggregate(total=Sum('quantity'))['total'] or 0
            total_losses = StockLoss.objects.filter(feed=feed).aggregate(total=Sum('quantity'))['total'] or 0
            total_used = Feeding.objects.filter(feed_type=feed).aggregate(total=Sum('quantity'))['total'] or 0

            available_stock = total_provisions - (total_used + total_losses)
            if feeding.quantity > available_stock:
                messages.error(
                    request,
                    f"Stock insuffisant pour {feed.name}. Disponible : {available_stock}, requis : {feeding.quantity}."
                )
            else:
                try:
                    feeding.full_clean()
                    feeding.save()
                    messages.success(request, "Feeding log added successfully.")
                    return redirect('feeding_list', batch_id=batch_id)
                except ValidationError as e:
                    for field, errors in e.message_dict.items():
                        for error in errors:
                            messages.error(request, f"{field}: {error}")
        else:
            messages.error(request, "There are some errors in the submitted data. Please check and try again.")
    else:
        form = FeedingForm(initial={'batch': batch_id})

    batch = get_object_or_404(Batch, id=batch_id)
    feeds = Feed.objects.all()

    context = {
        'grouped_feedings': page_obj,
        'form': form,
        'batch': batch,
        'feeds': feeds
    }

    return render(request, 'apps/feedings/list.html', context)


@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def feeding_list_print(request, batch_id):
    feedings = Feeding.objects.filter(batch_id=batch_id).order_by('feeding_date', 'id')
    query = request.GET.get('q')
    if query:
        feedings = feedings.filter(Q(feeding_date__icontains=query)).distinct().order_by('id')
    else:
        feedings = feedings
    company = CompanyProfile.objects.all().first()

    context = {
        'feedings': feedings,
        'company':company
    }

    return render(request, 'apps/feedings/list_print.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def feeding_list_export(request, batch_id):
    feedings = Feeding.objects.filter(batch_id=batch_id).order_by('feeding_date', 'id')
    query = request.GET.get('q')
    if query:
        feedings = feedings.filter(Q(feeding_date__icontains=query)).distinct().order_by('id')
    else:
        feedings = feedings

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedings"

    # Ajouter les en-têtes
    ws.append(["Date", "Batch", "Quantity", "Feed Type","Feed Stock Before","Feed Stock After", "Details"])

    # Ajouter les données des alimentations
    for feeding in feedings:
        ws.append([feeding.feeding_date, feeding.batch.name, feeding.quantity, feeding.feed_type.name,feeding.feed_quantity_before,feed_quantity_after, feeding.details])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Feedings.xlsx'

    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def update_feeding(request, id):
    feeding = get_object_or_404(Feeding, id=id)
    batch_id = feeding.batch.id

    if request.method == 'POST':
        # Mise à jour des champs avec les données du formulaire
        feeding_date = request.POST.get('feeding_date')
        quantity = request.POST.get('quantity')
        feed_type_id = request.POST.get('feed_type')
        details = request.POST.get('details')

        # Validation des données reçues
        if feeding_date:
            try:
                feeding.feeding_date = datetime.strptime(feeding_date, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")

        feeding.quantity = float(quantity) if quantity else feeding.quantity
        feeding.feed_type = get_object_or_404(Feed, id=feed_type_id) if feed_type_id else feeding.feed_type
        feeding.details = details if details else feeding.details

        # Vérification des stocks disponibles pour le feed utilisé
        feed = feeding.feed_type
        total_provisions = Provision.objects.filter(feed=feed).aggregate(total=Sum('quantity'))['total'] or 0
        total_losses = StockLoss.objects.filter(feed=feed).aggregate(total=Sum('quantity'))['total'] or 0
        total_used = Feeding.objects.filter(feed_type=feed).exclude(id=feeding.id).aggregate(total=Sum('quantity'))['total'] or 0

        available_stock = total_provisions - (total_used + total_losses)
        if feeding.quantity > available_stock:
            messages.error(
                request,
                f"Stock insuffisant pour {feed.name}. Disponible : {available_stock}, requis : {feeding.quantity}."
            )
        else:
            try:
                # Validation complète
                feeding.full_clean()
                feeding.save()
                messages.success(request, "Feeding log updated successfully.")
                return redirect('feeding_list', batch_id=batch_id)
            except ValidationError as e:
                for field, errors in e.message_dict.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")

    # Charger les informations pour le formulaire en cas d'erreur ou requête GET
    feeds = Feed.objects.all()

    context = {
        'feeding': feeding,
        'feeds': feeds,
        'batch_id': batch_id,
    }
    return render(request, 'apps/feedings/update.html', context)




@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def delete_feeding(request, id):
    feeding = get_object_or_404(Feeding, id=id)
    batch_id = feeding.batch.id
    feeding.delete()
    return redirect('feeding_list', batch_id=batch_id)

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def bulk_delete_feeding(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            feeding_ids = data.get('feeding_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not feeding_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucune alimentation sélectionnée.'}, status=400)

            elif action == 'delete':
                # Supprimer les alimentations
                feedings = Feeding.objects.filter(id__in=feeding_ids)
                feedings_deleted, _ = feedings.delete()

                return JsonResponse({'status': 'success', 'message': f'{feedings_deleted} alimentation(s) supprimée(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)



# Vues pour les Collectes d'Œufs
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def egg_collection_list(request, batch_id):
    egg_collections = EggCollection.objects.filter(batch_id=batch_id).order_by('collection_date', 'id')
    query = request.GET.get('q')
    if query:
        try:
            # Support both single date and date range queries
            if 'to' in query:
                start_date, end_date = query.split('to')
                start_date = datetime.strptime(start_date.strip(), '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date.strip(), '%Y-%m-%d').date()
                egg_collections = egg_collections.filter(collection_date__range=(start_date, end_date))
            else:
                query_date = datetime.strptime(query.strip(), '%Y-%m-%d').date()
                egg_collections = egg_collections.filter(collection_date=query_date)
        except ValueError:
            egg_collections = EggCollection.objects.none()

    # Regrouper les collections par date
    grouped_collections = defaultdict(list)
    for egg_collection in egg_collections:
        if egg_collection.collection_date:
            collection_date_str = egg_collection.collection_date.strftime('%Y-%m-%d')
            grouped_collections[collection_date_str].append(egg_collection)

    for egg_collection in egg_collections:
        if egg_collection.collection_date:
            egg_collection.collection_date = egg_collection.collection_date.strftime('%Y-%m-%d')

    # Paginer les groupes de collections
    paginator = Paginator(list(grouped_collections.items()), 6)  # Show 6 groups per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    batch = get_object_or_404(Batch, id=batch_id)
    if request.method == 'POST':
        form = EggCollectionForm(request.POST)
        if form.is_valid():
            egg_collection = form.save(commit=False)
            egg_collection.batch_id = batch_id
            egg_collection.save()
            return redirect('egg_collection_list', batch_id=batch_id)
    else:
        form = EggCollectionForm(initial={'batch': batch_id})

    context = {
        'grouped_collections': page_obj,
        'form': form,
        'batch': batch,
    }

    return render(request, 'apps/egg_collections/list.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def egg_collection_list_print(request, batch_id):
    egg_collections = EggCollection.objects.filter(batch_id=batch_id).order_by('collection_date', 'id')
    query = request.GET.get('q')
    if query:
        egg_collections = egg_collections.filter(Q(details__icontains=query)).distinct().order_by('id')
    else:
        egg_collections = egg_collections
    company = CompanyProfile.objects.all().first()
    context = {
        'egg_collections': egg_collections,
        'company':company
    }

    return render(request, 'apps/egg_collections/list_print.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def egg_collection_list_export(request, batch_id):
    egg_collections = EggCollection.objects.filter(batch_id=batch_id).order_by('collection_date', 'id')
    query = request.GET.get('q')
    if query:
        egg_collections = egg_collections.filter(Q(details__icontains=query)).distinct().order_by('id')
    else:
        egg_collections = egg_collections

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Egg Collections"
    # Ajouter les en-têtes
    ws.append(["Date", "Batch", "Quantity", "Cracked", "Details"])
    # Ajouter les données des collections
    for egg_collection in egg_collections:
        ws.append([egg_collection.collection_date, egg_collection.batch.name, egg_collection.quantity, egg_collection.craked, egg_collection.details])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=EggCollections.xlsx'
    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def update_egg_collection(request, id):
    egg_collection = get_object_or_404(EggCollection, id=id)

    if request.method == "POST":
        egg_collection.collection_date = request.POST.get('collection_date', egg_collection.collection_date)
        egg_collection.quantity = request.POST.get('quantity', egg_collection.quantity)
        egg_collection.craked = request.POST.get('craked', egg_collection.craked)
        egg_collection.details = request.POST.get('details', egg_collection.details)

        egg_collection.save()
        return redirect('egg_collection_list', egg_collection.batch.id)

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def delete_egg_collection(request, id):
    egg_collection = get_object_or_404(EggCollection, id=id)
    egg_collection.delete()
    return redirect('egg_collection_list', egg_collection.batch.id)

@csrf_exempt
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','employee']) 
def bulk_delete_egg_collection(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            eggcollection_ids = data.get('eggcollection_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not eggcollection_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucune collection sélectionnée.'}, status=400)
            elif action == 'delete':
                # Supprimer les collections et leurs images
                eggcollections = EggCollection.objects.filter(id__in=eggcollection_ids)
                eggcollections_deleted, _ = eggcollections.delete()
                return JsonResponse({'status': 'success', 'message': f'{eggcollections_deleted} collection(s) supprimé(s).'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)
        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Vues pour les Historiques de Traitements
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','employee']) 
def treatment_history_list(request, batch_id):
    treatment_histories = TreatmentHistory.objects.filter(batch_id=batch_id).order_by('treatment_date', 'id')
    query = request.GET.get('q')
    if query:
        try:
            if 'to' in query:  # Recherche par intervalle de dates
                start_date, end_date = query.split('to')
                start_date = datetime.strptime(start_date.strip(), '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date.strip(), '%Y-%m-%d').date()
                treatment_histories = treatment_histories.filter(treatment_date__range=(start_date, end_date))
            else:  # Recherche par date unique
                query_date = datetime.strptime(query.strip(), '%Y-%m-%d').date()
                treatment_histories = treatment_histories.filter(treatment_date=query_date)
        except ValueError:
            messages.error(request, "Invalid search format. Please use a valid date or date range.")
            treatment_histories = TreatmentHistory.objects.none()

    grouped_histories = defaultdict(list)
    for history in treatment_histories:
        if history.treatment_date:
            treatment_date_str = history.treatment_date.strftime('%Y-%m-%d')
            history.treatment_date = history.treatment_date.strftime('%Y-%m-%d')
            grouped_histories[treatment_date_str].append(history)

    paginator = Paginator(list(grouped_histories.items()), 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = TreatmentHistoryForm(request.POST)
        if form.is_valid():
            treatment_history = form.save(commit=False)
            treatment_history.batch_id = batch_id
            batch = get_object_or_404(Batch, id=batch_id)

            # Vérification des traitements existants
            existing_history = TreatmentHistory.objects.filter(
                batch_id=batch_id,
                treatment_date=treatment_history.treatment_date
            ).exists()

            if existing_history:
                messages.error(request, "A treatment for this date already exists.")
            else:
                treatment_history.save()
                messages.success(request, "Treatment history successfully added.")
                return redirect('treatment_history_list', batch_id=batch_id)
        else:
            messages.error(request, "Failed to add treatment history. Please check the form and try again.")

    else:
        form = TreatmentHistoryForm(initial={'batch': batch_id})

    batch = get_object_or_404(Batch, id=batch_id)
    treatments = Treatment.objects.all()

    context = {
        'grouped_histories': page_obj,
        'form': form,
        'batch': batch,
        'treatments': treatments
    }

    return render(request, 'apps/treatment_histories/list.html', context)


@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','employee']) 
def treatment_history_list_print(request, batch_id):
    treatment_histories = TreatmentHistory.objects.filter(batch_id=batch_id).order_by('treatment_date', 'id')
    query = request.GET.get('q')
    if query:
        treatment_histories = treatment_histories.filter(Q(treatment__name__icontains=query)).distinct().order_by('id')
    else:
        treatment_histories = treatment_histories
    company = CompanyProfile.objects.all().first()

    context = {
        'treatment_histories': treatment_histories,
        'company':company
    }

    return render(request, 'apps/treatment_histories/list_print.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','employee']) 
def treatment_history_list_export(request, batch_id):
    treatment_histories = TreatmentHistory.objects.filter(batch_id=batch_id).order_by('treatment_date', 'id')
    query = request.GET.get('q')
    if query:
        treatment_histories = treatment_histories.filter(Q(treatment__name__icontains=query)).distinct().order_by('id')
    else:
        treatment_histories = treatment_histories

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "TreatmentHistories"

    # Ajouter les en-têtes
    ws.append(["Date", "Batch", "Treatment", "Details"])

    # Ajouter les données des traitements
    for history in treatment_histories:
        ws.append([history.treatment_date, history.batch.name, history.treatment.name, history.details])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=TreatmentHistories.xlsx'

    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','employee']) 
def update_treatment_history(request, id):
    treatment_history = get_object_or_404(TreatmentHistory, id=id)

    if request.method == "POST":
        treatment_history.treatment = get_object_or_404(Treatment, id=request.POST.get('treatment'))
        treatment_history.treatment_date = request.POST.get('treatment_date', treatment_history.treatment_date)
        treatment_history.details = request.POST.get('details', treatment_history.details)
        batch_id = treatment_history.batch.id
        treatment_history.save()
        return redirect('treatment_history_list', batch_id=batch_id)

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def delete_treatment_history(request, id):
    treatment_history = get_object_or_404(TreatmentHistory, id=id)
    batch_id = treatment_history.batch.id
    treatment_history.delete()
    return redirect('treatment_history_list', batch_id=batch_id)

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def bulk_delete_treatment_history(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            treatment_history_ids = data.get('treatment_history_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not treatment_history_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucun traitement sélectionné.'}, status=400)

            elif action == 'delete':
                # Supprimer les traitements
                treatment_histories = TreatmentHistory.objects.filter(id__in=treatment_history_ids)
                treatment_histories_deleted, _ = treatment_histories.delete()

                return JsonResponse({'status': 'success', 'message': f'{treatment_histories_deleted} traitement(s) supprimé(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Vues pour les Provisions

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def provision_list(request):
    query = request.GET.get('q')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    provisions_query = Q()

    if query:
        provisions_query &= Q(feed__name__icontains=query)

    if date_from and date_to:
        provisions_query &= Q(provision_date__range=[date_from, date_to])
    elif date_from:
        provisions_query &= Q(provision_date__gte=date_from)
    elif date_to:
        provisions_query &= Q(provision_date__lte=date_to)

    provisions = Provision.objects.filter(provisions_query).distinct().order_by('provision_date')

    for provision in provisions:
        if provision.provision_date:
            provision.provision_date = provision.provision_date.strftime('%Y-%m-%d')
    paginator = Paginator(provisions, 10)  # Show 10 provisions per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = ProvisionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('provision_list')
    else:
        form = ProvisionForm()
    suppliers = Fournisseur.objects.all()
    feeds = Feed.objects.all()
    context = {
        'provisions': page_obj,
        'form': form,
        'suppliers': suppliers,
        'feeds': feeds,
    }

    return render(request, 'apps/provisions/list.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def provision_list_print(request):
    query = request.GET.get('q')
    if query:
        provisions = Provision.objects.filter(Q(feed__name__icontains=query)).distinct().order_by('id')
    else:
        provisions = Provision.objects.all().order_by('id')
    company = CompanyProfile.objects.all().first()
    context = {
        'provisions': provisions,
        'company':company
    }

    return render(request, 'apps/provisions/list_print.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def provision_list_export(request):
    query = request.GET.get('q')
    if query:
        provisions = Provision.objects.filter(Q(feed__name__icontains=query)).distinct().order_by('id')
    else:
        provisions = Provision.objects.all().order_by('id')

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Provisions"

    # Ajouter les en-têtes
    ws.append(["ID", "Supplier", "Feed", "Quantity", "Provision Date", "Details"])

    # Ajouter les données des provisions
    for provision in provisions:
        ws.append([provision.id, provision.supplier.name, provision.feed.name, provision.quantity, provision.provision_date, provision.details])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=provisions.xlsx'

    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def update_provision(request, id):
    provision = get_object_or_404(Provision, id=id)

    if request.method == "POST":
        form = ProvisionForm(request.POST, instance=provision)
        if form.is_valid():
            form.save()
            return redirect('provision_list')
        else:
            return render(request, 'apps/provisions/update_provision.html', {'form': form})

    form = ProvisionForm(instance=provision)
    return render(request, 'apps/provisions/update_provision.html', {'form': form})

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def delete_provision(request, id):
    provision = get_object_or_404(Provision, id=id)
    provision.delete()
    return redirect('provision_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def bulk_delete_provision(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            provision_ids = data.get('provision_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not provision_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucune provision sélectionnée.'}, status=400)

            elif action == 'delete':
                # Supprimer les provisions
                provisions = Provision.objects.filter(id__in=provision_ids)
                provisions_deleted, _ = provisions.delete()

                return JsonResponse({'status': 'success', 'message': f'{provisions_deleted} provision(s) supprimé(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)



#vues pour la perte des stock
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def stock_loss_list(request):
    query = request.GET.get('feed')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    losses_query = Q()

    if query:
        losses_query &= Q(feed__name__icontains=query)

    if date_from and date_to:
        losses_query &= Q(loss_date__range=[date_from, date_to])
    elif date_from:
        losses_query &= Q(loss_date__gte=date_from)
    elif date_to:
        losses_query &= Q(loss_date__lte=date_to)

    losses = StockLoss.objects.filter(losses_query).distinct().order_by('loss_date', 'id')

    for loss in losses:
        if loss.loss_date:
            loss.loss_date = loss.loss_date.strftime('%Y-%m-%d')

    paginator = Paginator(losses, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = StockLossForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock_loss_list')
    else:
        form = StockLossForm()

    

    context = {
        'losses': page_obj,
        'form': form,
    }
    return render(request, 'apps/stock_loss/list.html', context)


@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def stock_loss_list_print(request):
    losses = StockLoss.objects.all().order_by('loss_date', 'id')
    query = request.GET.get('q')
    if query:
        losses = losses.filter(Q(loss_date__icontains=query)).distinct().order_by('id')
    else:
        losses = StockLoss.objects.all().order_by('loss_date', 'id')
    company = CompanyProfile.objects.all().first()

    context = {
        'losses': losses,
        'company':company
    }

    return render(request, 'apps/stock_loss/list_print.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def stock_loss_export(request):
    losses = StockLoss.objects.all().order_by('loss_date', 'id')
    query = request.GET.get('q')
    if query:
        losses = losses.filter(Q(loss_date__icontains=query)).distinct().order_by('id')

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Losses"
    ws.append(["Date", "Feed", "Quantity", "Reason", "Details"])

    for loss in losses:
        ws.append([loss.loss_date, loss.feed.name, loss.quantity, loss.reason, loss.details])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=StockLosses.xlsx'
    wb.save(response)
    return response

@login_required(login_url=login_path)
def update_stock_loss(request, id):
    loss = get_object_or_404(StockLoss, id=id)
    if request.method == 'POST':
        form = StockLossForm(request.POST, instance=loss)
        if form.is_valid():
            form.save()
            return redirect('stock_loss_list')
    else:
        form = StockLossForm(instance=loss)
    return render(request, 'apps/stock_loss/update.html', {'form': form})

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def delete_stock_loss(request, id):
    loss = get_object_or_404(StockLoss, id=id)
    loss.delete()
    return redirect('stock_loss_list')


@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def bulk_delete_stock_loss(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            loss_ids = data.get('loss_ids', [])
            if not loss_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucune perte sélectionnée.'}, status=400)

            losses = StockLoss.objects.filter(id__in=loss_ids)
            deleted_count, _ = losses.delete()
            return JsonResponse({'status': 'success', 'message': f'{deleted_count} perte(s) supprimée(s).'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)


# Vues pour les Catégories de Charges
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def expense_category_list(request):
    query = request.GET.get('q')
    if query:
        categories = ExpenseCategory.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        categories = ExpenseCategory.objects.all().order_by('id')

    paginator = Paginator(categories, 10)  # Show 10 categories per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('expense_category_list')
    else:
        form = ExpenseCategoryForm()

    context = {
        'categories': page_obj,
        'form': form,
    }

    return render(request, 'apps/expense_categories/list.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def expense_category_list_print(request):
    query = request.GET.get('q')
    if query:
        categories = ExpenseCategory.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        categories = ExpenseCategory.objects.all().order_by('id')
    company = CompanyProfile.objects.all().first()

    context = {
        'categories': categories,
        'company':company
    }

    return render(request, 'apps/expense_categories/list_print.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def expense_category_list_export(request):
    query = request.GET.get('q')
    if query:
        categories = ExpenseCategory.objects.filter(Q(name__icontains=query)).distinct().order_by('id')
    else:
        categories = ExpenseCategory.objects.all().order_by('id')

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Categories"

    # Ajouter les en-têtes
    ws.append(["ID", "Name"])

    # Ajouter les données des catégories
    for category in categories:
        ws.append([category.id, category.name])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=expense_categories.xlsx'

    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def update_expense_category(request, id):
    category = get_object_or_404(ExpenseCategory, id=id)

    if request.method == "POST":
        category.name = request.POST.get('name', category.name)

        try:
            category.save()
            print("ExpenseCategory saved successfully")
        except Exception as e:
            print(f"Error saving expense category: {e}")

        return redirect('expense_category_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def delete_expense_category(request, id):
    category = get_object_or_404(ExpenseCategory, id=id)
    category.delete()
    return redirect('expense_category_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def bulk_delete_expense_categories(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            category_ids = data.get('category_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not category_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucune catégorie sélectionnée.'}, status=400)

            elif action == 'delete':
                # Supprimer les catégories
                categories = ExpenseCategory.objects.filter(id__in=category_ids)
                categories_deleted, _ = categories.delete()

                return JsonResponse({'status': 'success', 'message': f'{categories_deleted} catégorie(s) supprimée(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Vues pour les Charges
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def expense_list(request):
    query = request.GET.get('q')
    if query:
        expenses = Expense.objects.filter(Q(description__icontains=query) | Q(category__name__icontains=query)).distinct().order_by('id')
    else:
        expenses = Expense.objects.all().order_by('id')

    paginator = Paginator(expenses, 10)  # Show 10 expenses per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('expense_list')
        else:
            print(form.errors) 
    else:
        form = ExpenseForm()

    context = {
        'expenses': page_obj,
        'form': form,
    }

    return render(request, 'apps/expenses/list.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def expense_list_print(request):
    query = request.GET.get('q')
    if query:
        expenses = Expense.objects.filter(Q(description__icontains=query) | Q(category__name__icontains=query)).distinct().order_by('id')
    else:
        expenses = Expense.objects.all().order_by('id')
    company = CompanyProfile.objects.all().first()

    context = {
        'expenses': expenses,
        'company':company
    }

    return render(request, 'apps/expenses/list_print.html', context)

@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def expense_list_export(request):
    query = request.GET.get('q')
    if query:
        expenses = Expense.objects.filter(Q(description__icontains=query) | Q(category__name__icontains=query)).distinct().order_by('id')
    else:
        expenses = Expense.objects.all().order_by('id')

    # Créer un nouveau classeur
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Ajouter les en-têtes
    ws.append(["ID", "Category", "Supplier", "Batch", "Description", "Amount", "Paid Amount", "Expense Date", "Debt Amount"])

    # Ajouter les données des charges
    for expense in expenses:
        ws.append([expense.id, expense.category.name, expense.supplier.name if expense.supplier else "", expense.batch.name if expense.batch else "", expense.description, expense.amount, expense.paid_amount, expense.expense_date, expense.debt_amount])

    # Créer une réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'

    # Enregistrer le classeur dans la réponse
    wb.save(response)

    return response

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def update_expense(request, id):
    expense = get_object_or_404(Expense, id=id)

    if request.method == "POST":
        expense.category = get_object_or_404(ExpenseCategory, id=request.POST.get('category'))
        expense.supplier = get_object_or_404(Fournisseur, id=request.POST.get('supplier')) if request.POST.get('supplier') else None
        expense.batch = get_object_or_404(Batch, id=request.POST.get('batch')) if request.POST.get('batch') else None
        expense.description = request.POST.get('description', expense.description)
        expense.amount = request.POST.get('amount', expense.amount)
        expense.paid_amount = request.POST.get('paid_amount', expense.paid_amount)
        expense.expense_date = request.POST.get('expense_date', expense.expense_date)

        try:
            expense.save()
            print("Expense saved successfully")
        except Exception as e:
            print(f"Error saving expense: {e}")

        return redirect('expense_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def delete_expense(request, id):
    expense = get_object_or_404(Expense, id=id)
    expense.delete()
    return redirect('expense_list')

@csrf_exempt
@login_required(login_url=login_path)
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def bulk_delete_expenses(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON envoyées
            data = json.loads(request.body)
            expense_ids = data.get('expense_ids', [])
            action = data.get('action', 'delete')  # Action à effectuer: 'toggle' ou 'delete'

            if not expense_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucune charge sélectionnée.'}, status=400)

            elif action == 'delete':
                # Supprimer les charges
                expenses = Expense.objects.filter(id__in=expense_ids)
                expenses_deleted, _ = expenses.delete()

                return JsonResponse({'status': 'success', 'message': f'{expenses_deleted} charge(s) supprimée(s).'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Action non reconnue.'}, status=400)

        except Exception as e:
            # Renvoie une erreur avec un message détaillé
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)


#allocations

@login_required(login_url='login_path')
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def batch_allocation_list(request, batch_id):
    allocations = BatchAllocation.objects.filter(batch_id=batch_id).order_by('id')
    allocation_id = request.GET.get('allocation_id')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    allocations_query = Q()

    if allocation_id:
        allocations_query &= Q(id=allocation_id)

    if date_from and date_to:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            allocations_query &= Q(order_item__order__order_date__range=(start_date, end_date))
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
            allocations = BatchAllocation.objects.none()
    elif date_from:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            allocations_query &= Q(order_item__order__order_date__gte=start_date)
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
            allocations = BatchAllocation.objects.none()
    elif date_to:
        try:
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            allocations_query &= Q(order_item__order__order_date__lte=end_date)
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
            allocations = BatchAllocation.objects.none()

    allocations = allocations.filter(allocations_query).distinct().order_by('id')

    paginator = Paginator(allocations, 10)  # 10 allocations per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    batch = get_object_or_404(Batch, id=batch_id)

   

    context = {
        'allocations': page_obj,
        'batch': batch,
        
    }

    return render(request, 'apps/batch_allocations/list.html', context)

@csrf_exempt
@login_required(login_url='login_path')
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def delete_batch_allocation(request, id):
    allocation = get_object_or_404(BatchAllocation, id=id)
    batch_id = allocation.batch.id
    allocation.delete()
    return redirect('batch_allocation_list', batch_id=batch_id)

@csrf_exempt
@login_required(login_url='login_path')
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def bulk_delete_batch_allocation(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            allocation_ids = data.get('allocation_ids', [])
            action = data.get('action', 'delete')

            if not allocation_ids:
                return JsonResponse({'status': 'error', 'message': 'No allocations selected.'}, status=400)

            elif action == 'delete':
                allocations = BatchAllocation.objects.filter(id__in=allocation_ids)
                allocations_deleted, _ = allocations.delete()

                return JsonResponse({'status': 'success', 'message': f'{allocations_deleted} allocation(s) deleted.'})

            else:
                return JsonResponse({'status': 'error', 'message': 'Unrecognized action.'}, status=400)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Method not allowed.'}, status=405)

@login_required(login_url='login_path')
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def batch_allocation_list_export(request, batch_id):
    allocations = BatchAllocation.objects.filter(batch_id=batch_id).order_by('id')
    query = request.GET.get('q')
    if query:
        allocations = allocations.filter(Q(order_item__order__order_date__icontains=query)).distinct().order_by('id')
    else:
        allocations = allocations

    wb = Workbook()
    ws = wb.active
    ws.title = "BatchAllocations"

    ws.append(["Order Item", "Batch", "Quantity Eggs", "Quantity Poultry"])

    for allocation in allocations:
        ws.append([allocation.order_item, allocation.batch.name, allocation.quantity_eggs, allocation.quantity_poultry])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=BatchAllocations.xlsx'

    wb.save(response)

    return response

@login_required(login_url='login_path')
@role_required(excluded_roles=['customer','cashier','veterinarian','employee','farmer']) 
def batch_allocation_list_print(request, batch_id):
    allocations = BatchAllocation.objects.filter(batch_id=batch_id).order_by('id')
    query = request.GET.get('q')
    if query:
        allocations = allocations.filter(Q(order_item__order__order_date__icontains=query)).distinct().order_by('id')
    else:
        allocations = allocations
    company = CompanyProfile.objects.all().first()
    context = {
        'allocations': allocations,
        'company':company
    }

    return render(request, 'apps/batch_allocations/list_print.html', context)
    allocations = BatchAllocation.objects.filter(batch_id=batch_id).order_by('id')
    query = request.GET.get('q')

    if query:
        try:
            if 'to' in query:
                start_date, end_date = query.split('to')
                start_date = datetime.strptime(start_date.strip(), '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date.strip(), '%Y-%m-%d').date()
                allocations = allocations.filter(order_item__order__order_date__range=(start_date, end_date))
            else:
                query_date = datetime.strptime(query.strip(), '%Y-%m-%d').date()
                allocations = allocations.filter(order_item__order__order_date=query_date)
        except ValueError:
            allocations = BatchAllocation.objects.none()
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")

    paginator = Paginator(allocations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    batch = get_object_or_404(Batch, id=batch_id)

    if request.method == 'POST':
        form = BatchAllocationForm(request.POST)
        if form.is_valid():
            allocation = form.save(commit=False)
            allocation.batch_id = batch_id
            allocation.save()
            return redirect('batch_allocation_list', batch_id=batch_id)
        else:
            messages.error(request, "There are some errors in the submitted data. Please check and try again.")
    form = BatchAllocationForm()
    context = {
        'allocations': page_obj,
        'batch': batch,
        'form': form,
    }

    return render(request, 'apps/batch_allocations/list.html', context)



@login_required(login_url='login_path')
def batch_statistics_view(request):
    # Récupérer les paramètres de la requête
    from_date = request.GET.get('from')
    to_date = request.GET.get('to')
    building_id = request.GET.get('building_id')

    # Filtrer les lots par date et bâtiment
    query = Q()
    if from_date:
        query &= Q(arrival_date__gte=from_date)
    if to_date:
        query &= Q(arrival_date__lte=to_date)
    if building_id:
        query &= Q(building_id=building_id)

    batches = Batch.objects.filter(query)

    # Préparer les données à renvoyer
    data = []
    for batch in batches:
        batch_data = {
            'id':batch.id,
            'name': batch.name,
            'breed': batch.breed.name,
            'building': batch.building.name,
            'arrival_date': batch.arrival_date,
            'arrival_age': batch.arrival_age,
            'arrival_quantity': batch.arrival_quantity,
            'status': batch.status,
            'details': batch.details,
            'current_age': batch.get_current_age(),
            'deceased_quantity': batch.get_deceased_quantity(),
            'sick_quantity': batch.get_sick_quantity(),
            'current_poultry': batch.get_current_poultry(),
            'available_poultry': batch.get_available_poultry(),
            'total_eggs_collected': batch.get_total_eggs_collected(),
            'available_eggs': batch.get_available_eggs(),
            'revenue_from_eggs': batch.get_revenue_from_eggs(),
            'revenue_from_poultry': batch.get_revenue_from_poultry(),
            'total_revenue': batch.get_total_revenue(),
            'total_expenses': batch.get_total_expenses(),
            'mortality_rate': batch.get_mortality_rate(),
            'laying_rate': batch.get_laying_rate(),
            'nutrition_data': batch.get_nutrition_data(),
            'monthly_profitability': batch.calculate_monthly_profitability(),
        }
        data.append(batch_data)

    return JsonResponse(data, safe=False)
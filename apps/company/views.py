import os
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from openpyxl import Workbook
from django.conf import settings
from kake.views.mixins import TenantCreationMixin
from .models import Company, Subscription, Domain, Payment
from .forms import CompanyForm, SubscriptionForm, PaymentForm  # Ensure you have these forms
from apps.users.decorators import role_required

# Liste des compagnies
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def company_list(request):
    query = request.GET.get('q')
    companies = Company.objects.filter(Q(name__icontains=query)).distinct().order_by('id') if query else Company.objects.all().order_by('id')

    paginator = Paginator(companies, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Gestion du formulaire de création
    if request.method == 'POST':
        form = CompanyForm(request.POST)
        if form.is_valid():
            try:
                company = form.save()
                print(request.user.role)

                # Créer et migrer le tenant
                tenant_creation_mixin = TenantCreationMixin()
                tenant_creation_mixin.create_and_migrate_tenant(
                    name=company.name,
                    schema_name=company.name.lower().replace(' ', '_'),
                    domain=company.name.lower().replace(' ', '-') + '.localhost:8000',
                    user=request.user  # Transmettre l'utilisateur ici
                )

                messages.success(request, "La compagnie a été créée avec succès.")
                return redirect('company_list')
            except Exception as e:
                company.delete()
                messages.error(request, f"Erreur lors de la création du tenant : {str(e)}")
        else:
            messages.error(request, "Erreur dans les données du formulaire.")
    else:
        form = CompanyForm()

    return render(request, 'apps/companies/list.html', {'companies': page_obj, 'form': form})

# Vue pour impression
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def company_list_print(request):
    query = request.GET.get('q')
    companies = Company.objects.filter(Q(name__icontains=query)).distinct() if query else Company.objects.all()
    return render(request, 'apps/companies/list_print.html', {'companies': companies})

# Export des compagnies
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def company_list_export(request):
    query = request.GET.get('q')
    companies = Company.objects.filter(Q(name__icontains=query)).distinct() if query else Company.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    ws.append(["ID", "Name", "Email", "Phone", "Subscription Plan", "Active"])

    for company in companies:
        ws.append([company.id, company.name, company.email, company.phone, company.subscription_plan, company.is_active])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=companies.xlsx'
    wb.save(response)

    return response

# Mise à jour d'une compagnie
@csrf_exempt
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def update_company(request, id):
    company = get_object_or_404(Company, id=id)
    if request.method == "POST":
        company.name = request.POST.get('name', company.name)
        company.email = request.POST.get('email', company.email)
        company.phone = request.POST.get('phone', company.phone)
        company.subscription_plan = request.POST.get('subscription_plan', company.subscription_plan)
        company.is_active = request.POST.get('is_active', company.is_active) == 'true'

        company.save()
        messages.success(request, "La compagnie a été mise à jour avec succès.")
        return redirect('company_list')

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Suppression d'une compagnie
@csrf_exempt
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def delete_company(request, id):
    company = get_object_or_404(Company, id=id)
    company.delete()
    messages.success(request, "La compagnie a été supprimée avec succès.")
    return redirect('company_list')

# Suppression en masse
@csrf_exempt
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def bulk_delete_companies(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            company_ids = data.get('company_ids', [])

            if not company_ids:
                return JsonResponse({'status': 'error', 'message': 'Aucune entreprise sélectionnée.'}, status=400)

            companies_deleted, _ = Company.objects.filter(id__in=company_ids).delete()
            return JsonResponse({'status': 'success', 'message': f'{companies_deleted} entreprise(s) supprimée(s).'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)



# Liste des abonnements
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def subscription_list(request):
    query = request.GET.get('q')
    subscriptions = Subscription.objects.filter(Q(company__name__icontains=query)).distinct().order_by('-start_date') if query else Subscription.objects.all().order_by('-start_date')

    paginator = Paginator(subscriptions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Handle the creation form
    if request.method == 'POST':
        form = SubscriptionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "L'abonnement a été créé avec succès.")
            return redirect('subscription_list')
        else:
            messages.error(request, "Erreur dans les données du formulaire.")
    else:
        form = SubscriptionForm()

    return render(request, 'apps/subscriptions/list.html', {'subscriptions': page_obj, 'form': form})

# Vue pour impression
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def subscription_list_print(request):
    query = request.GET.get('q')
    subscriptions = Subscription.objects.filter(Q(company__name__icontains=query)).distinct() if query else Subscription.objects.all()
    return render(request, 'apps/subscriptions/list_print.html', {'subscriptions': subscriptions})

# Export des abonnements
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def subscription_list_export(request):
    query = request.GET.get('q')
    subscriptions = Subscription.objects.filter(Q(company__name__icontains=query)).distinct() if query else Subscription.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Subscriptions"
    ws.append(["ID", "Company", "Plan", "Start Date", "End Date", "Active"])

    for sub in subscriptions:
        ws.append([sub.id, sub.company.name, sub.get_plan_display(), sub.start_date, sub.end_date, sub.active])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=subscriptions.xlsx'
    wb.save(response)

    return response

# Mise à jour d'un abonnement
@csrf_exempt
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def update_subscription(request, id):
    subscription = get_object_or_404(Subscription, id=id)
    if request.method == "POST":
        subscription.plan = request.POST.get('plan', subscription.plan)
        subscription.end_date = request.POST.get('end_date', subscription.end_date)
        subscription.active = request.POST.get('active', subscription.active) == 'true'
        subscription.save()

        messages.success(request, "L'abonnement a été mis à jour avec succès.")
        return redirect('subscription_list')

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Suppression d'un abonnement
@csrf_exempt
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def delete_subscription(request, id):
    subscription = get_object_or_404(Subscription, id=id)
    subscription.delete()
    messages.success(request, "L'abonnement a été supprimé avec succès.")
    return redirect('subscription_list')



# Liste des paiements
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def payment_list(request):
    query = request.GET.get('q')
    payments = Payment.objects.filter(Q(subscription__company__name__icontains=query)).distinct().order_by('-payment_date') if query else Payment.objects.all().order_by('-payment_date')

    paginator = Paginator(payments, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Handle the creation form
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Le paiement a été créé avec succès.")
            return redirect('payment_list')
        else:
            messages.error(request, "Erreur dans les données du formulaire.")
    else:
        form = PaymentForm()

    return render(request, 'apps/payments/list.html', {'payments': page_obj, 'form': form})

# Vue pour impression
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def payment_list_print(request):
    query = request.GET.get('q')
    payments = Payment.objects.filter(Q(subscription__company__name__icontains=query)).distinct() if query else Payment.objects.all()
    return render(request, 'apps/payments/list_print.html', {'payments': payments})

# Export des paiements
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def payment_list_export(request):
    query = request.GET.get('q')
    payments = Payment.objects.filter(Q(subscription__company__name__icontains=query)).distinct() if query else Payment.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    ws.append(["ID", "Subscription", "Amount", "Payment Status", "Payment Method", "Transaction ID", "Payment Date", "Remark"])

    for payment in payments:
        ws.append([payment.id, payment.subscription.id, payment.amount, payment.get_payment_status_display(), payment.get_payment_method_display(), payment.transaction_id, payment.payment_date, payment.remark])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=payments.xlsx'
    wb.save(response)

    return response

# Mise à jour d'un paiement
@csrf_exempt
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def update_payment(request, id):
    payment = get_object_or_404(Payment, id=id)
    if request.method == "POST":
        payment.amount = request.POST.get('amount', payment.amount)
        payment.payment_status = request.POST.get('payment_status', payment.payment_status)
        payment.payment_method = request.POST.get('payment_method', payment.payment_method)
        payment.transaction_id = request.POST.get('transaction_id', payment.transaction_id)
        payment.remark = request.POST.get('remark', payment.remark)
        payment.save()

        messages.success(request, "Le paiement a été mis à jour avec succès.")
        return redirect('payment_list')

    return JsonResponse({'status': 'error', 'message': 'Méthode non autorisée.'}, status=405)

# Suppression d'un paiement
@csrf_exempt
@login_required(login_url='login')
@role_required(excluded_roles=['customer','cashier','veterinarian','accountant','employee','farmer']) 
def delete_payment(request, id):
    payment = get_object_or_404(Payment, id=id)
    payment.delete()
    messages.success(request, "Le paiement a été supprimé avec succès.")
    return redirect('payment_list')
